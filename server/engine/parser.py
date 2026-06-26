"""
parser.py — Shakl 1 (Balans) va Shakl 2 (Moliyaviy natijalar) ni
QATOR KODI bo'yicha o'qiydi.

  Shakl 1: 010–780 kodlar (Aktiv: 010–400, Passiv: 410–780)
  Shakl 2: 010–270 kodlar

Ishlash tartibi:
  1. Har bir varaq Shakl 1 yoki Shakl 2 ga ajratiladi (matn + kodlar bo'yicha).
  2. Varaqdagi "satr kodi" ustuni avtomatik aniqlanadi.
  3. Kerakli kodlar qatoridan qiymat olinadi (oxirgi 2 son = boshi/oxiri yoki
     o'tgan/joriy; oxirgisi joriy deb olinadi).
  4. Kod topilmasa — kalit so'z bo'yicha qidiriladi (zaxira usul).

MUHIM: kod raqamlari НСБУ shakl versiyasiga qarab biroz farq qilishi mumkin.
Quyidagi xaritalar klassik (010–780 / 010–270) versiyaga mos. Agar biror
ko'rsatkich noto'g'ri chiqsa — FORM1_CODES / FORM2_CODES dagi raqamni
faylingizdagi "satr kodi" ustuniga qarab to'g'rilang.
"""

from __future__ import annotations
import re
from openpyxl import load_workbook
from indicators import FinancialData


class FormParseError(Exception):
    """Mijozga ko'rsatish mumkin bo'lgan, tushunarli o'qish xatosi.

    code   : xato turi ('missing_fields', 'empty', 'no_codes' ...)
    message: o'zbekcha tushunarli matn (mijozga ko'rsatish uchun)
    fields : yetishmayotgan ko'rsatkichlar ro'yxati (ixtiyoriy)
    """
    def __init__(self, code: str, message: str, fields=None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.fields = fields or []

# ---------------------------------------------------------------------------
# KOD XARITALARI  (kod -> FinancialData maydoni)
# ---------------------------------------------------------------------------
# Shakl 1 — Balans. Bizga faqat jami/yakuniy qatorlar kerak.
FORM1_CODES = {
    130: "non_current_assets",      # Итого по разделу I (Uzoq muddatli aktivlar)
    140: "inventories",             # Товарно-материальные запасы
    210: "receivables",             # Дебиторы, всего
    320: "cash",                    # Денежные средства, всего
    370: "short_term_investments",  # Краткосрочные инвестиции
    390: "current_assets",          # Итого по разделу II (Joriy aktivlar)
    400: "total_assets",            # Всего по активу баланса
    480: "equity",                  # Итого по разделу I пассива (kapital)
    490: "long_term_liabilities",   # Долгосрочные обязательства, всего
    600: "current_liabilities",     # Текущие обязательства, всего
    # 770 = Итого по разделу II (barcha majburiyatlar) — kerak bo'lsa
}

# Shakl 1 — SUG'URTA tashkilotlari uchun maxsus shakl (kodlar farq qiladi!)
# Aniqlash belgisi: varaqda 1200 (passiv jami) yoki 930+570 kodlari bor.
FORM1_INS_CODES = {
    130: "non_current_assets",      # Итого по разделу I
    140: "inventories",             # Товарно-материальные запасы, всего
    190: "receivables",             # Дебиторы, всего
    410: "cash",                    # Денежные средства, всего
    460: "short_term_investments",  # Краткосрочные инвестиции
    480: "current_assets",          # Итого по разделу II
    490: "total_assets",            # Всего по активу баланса
    570: "equity",                  # Итого по разделу I пассива (kapital)
    730: "long_term_liabilities",   # Долгосрочные обязательства, всего
    930: "current_liabilities",     # Текущие обязательства, всего
}
INS_RESERVES_CODE = 720             # Sug'urta rezervlari (sof) — majburiyatga qo'shiladi

# Shakl 2 — Moliyaviy natijalar (standart НСБУ kodlari).
# Versiyaga qarab farq qilishi mumkin — kalit so'z zaxirasi ham ishlaydi.
FORM2_CODES = {
    10:  "revenue",            # Sotuvdan sof tushum / Чистая выручка
    20:  "cogs",               # Tannarx / Себестоимость
    30:  "gross_profit",       # Yalpi foyda / Валовая прибыль
    40:  "period_expenses",    # Davr xarajatlari / Расходы периода
    100: "operating_profit",   # Прибыль от основной деятельности
    240: "profit_before_tax",  # Прибыль до уплаты налога на прибыль
    250: "income_tax",         # Налог на прибыль
    270: "net_profit",         # Чистая прибыль отчётного периода
}

# Zaxira: kalit so'zlar (kod topilmaganda)
KEYWORDS = {
    "non_current_assets": ["uzoq muddatli aktivlar", "i bo'lim bo'yicha jami", "долгосрочные активы"],
    "current_assets":     ["joriy aktivlar", "ii bo'lim bo'yicha jami", "оборотные активы", "текущие активы"],
    "total_assets":       ["balans aktivi", "jami aktiv", "aktiv balansi", "итого актив", "всего актив"],
    "inventories":        ["tovar-moddiy zaxira", "zaxira", "запас"],
    "receivables":        ["debitorlik", "дебиторск"],
    "cash":               ["pul mablag", "naqd pul", "денежные средства"],
    "short_term_investments": ["qisqa muddatli investitsiya", "краткосрочные инвестиции"],
    "equity":             ["o'z mablag'lari manbalari", "kapital jami", "o'z kapitali", "собственн", "итого капитал"],
    "long_term_liabilities": ["uzoq muddatli majburiyat", "долгосрочные обязательства"],
    "current_liabilities":["joriy majburiyat", "qisqa muddatli majburiyat", "текущие обязательства", "краткосрочные обязательства"],
    "revenue":            ["sotishdan sof tushum", "sof tushum", "выручка"],
    "cogs":               ["sotilgan mahsulot tannarxi", "tannarx", "себестоимост"],
    "gross_profit":       ["yalpi foyda", "валовая прибыль"],
    "period_expenses":    ["davr xarajatlari", "расходы периода"],
    "operating_profit":   ["asosiy faoliyat", "основной деятельности"],
    "profit_before_tax":  ["soliq to'lashgacha", "soliqqacha", "до уплаты налог", "до налог"],
    "income_tax":         ["foyda solig", "налог на прибыль"],
    "net_profit":         ["sof foyda", "hisobot davri sof", "чистая прибыль"],
}


# ---------------------------------------------------------------------------
# Yordamchi funksiyalar
# ---------------------------------------------------------------------------
def _to_float(x):
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[()]", "", s)  # qavslarni (manfiy) olib tashlaymiz
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _as_code(cell):
    """Hujayra qator kodimi? Ha bo'lsa int kod (10..1200, 10 ga karrali).
    Sug'urta shaklida kodlar 1200 gacha boradi."""
    if cell is None or isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)):
        v = int(cell)
        if v == cell and 10 <= v <= 1200 and v % 10 == 0:
            return v
        return None
    s = str(cell).strip()
    if re.fullmatch(r"0*\d{2,4}", s):
        v = int(s)
        if 10 <= v <= 1200 and v % 10 == 0:
            return v
    return None


def _detect_code_column(rows: list[tuple]) -> int | None:
    """Qaysi ustunda kodlar eng ko'p uchrashini topadi."""
    counts: dict[int, int] = {}
    for row in rows:
        for ci, cell in enumerate(row):
            if _as_code(cell) is not None:
                counts[ci] = counts.get(ci, 0) + 1
    if not counts:
        return None
    return max(counts, key=counts.get)


def _classify(sheet_text: str, codes_present: set[int]) -> str:
    """Varaq Shakl 1 (balans) yoki Shakl 2 (natijalar)?"""
    f2 = any(k in sheet_text for k in
             ["moliyaviy natija", "sof tushum", "sotishdan", "выручка", "финансовых результат"])
    f1 = (any(k in sheet_text for k in ["balans", "aktiv", "passiv", "баланс"])
          or {400, 480, 780} & codes_present)
    if f1 and not f2:
        return "form1"
    if f2 and not f1:
        return "form2"
    return "mixed"  # ikkalasi yoki noaniq


def _extract_codes(rows: list[tuple], code_col: int) -> dict[int, tuple]:
    """{kod: (prev, cur)} — kod ustunidan keyingi raqamlardan.
    Nol va bo'sh ('x') qiymatlar tashlanadi: Shakl 2 da ishlatilmagan
    daromad/xarajat ustunlarida 0 turishi mumkin, ular joriy davrni buzmasligi
    uchun. Shunday qilib qoladigan sonlar [oldingi_davr, joriy_davr] bo'ladi."""
    out: dict[int, tuple] = {}
    for row in rows:
        if code_col >= len(row):
            continue
        code = _as_code(row[code_col])
        if code is None:
            continue
        all_nums = [v for v in (_to_float(c) for c in row[code_col + 1:])
                    if v is not None]
        nonzero = [v for v in all_nums if v != 0]
        if nonzero:
            cur = nonzero[-1]
            prev = nonzero[-2] if len(nonzero) >= 2 else None
        elif all_nums:
            # qator bor, lekin hammasi nol -> qiymat haqiqatan 0
            cur = 0.0
            prev = 0.0 if len(all_nums) >= 2 else None
        else:
            continue
        out.setdefault(code, (prev, cur))
    return out


# ---------------------------------------------------------------------------
# Asosiy funksiya
# ---------------------------------------------------------------------------
def _cell_text(cell) -> str:
    return "" if cell is None else str(cell).strip()


def _next_value_after_label(cells: list, label_idx: int, label_terms: tuple[str, ...]) -> str:
    """Label turgan katakdan keyingi mazmunli qiymatni oladi."""
    for c in cells[label_idx + 1:]:
        t = _cell_text(c)
        tl = t.lower()
        if not t or t.endswith(":"):
            continue
        if any(term in tl for term in label_terms):
            continue
        if tl in ("код", "код строки", "сатр коди", "satr kodi", "code"):
            continue
        if t.replace(".", "").replace("-", "").isdigit():
            continue
        return t
    return ""


def _extract_header(rows: list[tuple]) -> dict:
    """Sarlavhadan korxona nomi, STIR (ИНН) va hisobot davrini o'qiydi.
    Standart НСБУ shakli ham, emitent yillik hisoboti ham qo'llanadi."""
    company = ""
    stir = ""
    year = None
    quarter = None
    for row in rows:
        cells = [c for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        joined = " ".join(str(c) for c in cells).lower()

        company_terms = ("предприят", "организац", "korxona", "tashkilot",
                         "корхона", "ташкилот", "полное", "эмитент",
                         "company")
        stir_terms = ("идентификационный номер", "инн", "стир", "stir",
                      "налогоплательщ", "налоговой")

        # Korxona nomi (standart: "Предприятие...", emitent: "Полное:")
        if not company and any(term in joined for term in company_terms):
            if "наименование показател" in joined or "кўрсаткичлар номи" in joined:
                continue
            if any(_as_code(c) is not None for c in cells):
                continue
            if re.search(r"\(\s*\d{3,4}\s*\)", joined):
                continue
            if any(bad in joined for bad in ("инвестици", "капитал", "актив", "пассив",
                                             "сатр", "строк", "показател", "кўрсаткич")):
                continue
            # "Орган эмитента, утвердивший...", "Дата..." kabi qatorlar nom emas
            # (diqqat: "орган" so'zini ishlatib bo'lmaydi — u "организация"
            #  ichida ham uchraydi)
            if any(bad in joined for bad in ("утверд", "дата", "адрес")):
                continue
            for i, c in enumerate(cells):
                if any(term in str(c).lower() for term in company_terms):
                    company = _next_value_after_label(cells, i, company_terms)
                    if company:
                        break
            for c in cells:
                if company:
                    break
                t = str(c).strip()
                tl = t.lower()
                if (len(t) > 6 and not t.endswith(":")
                        and not tl.startswith("по ")
                        and not any(k in tl for k in ("предприят", "организац",
                                                      "korxona", "tashkilot",
                                                      "полное", "эмитент",
                                                      "наименование"))
                        and not t.replace(".", "").replace("-", "").isdigit()):
                    company = t
                    break

        # STIR / ИНН (9 xonali raqam)
        if not stir and any(term in joined for term in stir_terms):
            for c in cells:
                d = re.sub(r"\D", "", str(c))
                if len(d) == 9:
                    stir = d
                    break

        # Davr: yil + chorak (standart shakl)
        if (year is None) and ("квартал" in joined or "chorak" in joined
                               or "чорак" in joined or "чораг" in joined
                               or "год" in joined or "yil" in joined
                               or "йил" in joined):
            for c in cells:
                s = str(c).strip()
                if re.fullmatch(r"20\d{2}", s):
                    year = int(s)
                    break
            for c in cells:
                s = str(c).strip()
                if re.fullmatch(r"[1-4]", s):
                    quarter = int(s)
                    break

        # Davr: "Дата отчетности: 2025-12-31" (emitent hisoboti)
        if year is None and "дата отчетности" in joined:
            m = re.search(r"(20\d{2})", joined)
            if m:
                year = int(m.group(1))

    return {"company": company, "stir": stir, "year": year, "quarter": quarter}


def _load_all_sheets(path: str) -> list[list[tuple]]:
    """Excel fayldan barcha varaqlarning satrlarini o'qiydi.
    Formatni FAYL ICHIDAGI magic-baytlardan aniqlaydi (nomga ishonmaymiz,
    chunki bot hamma faylni .xlsx nomi bilan saqlaydi):
      PK..   -> XLSX (openpyxl)
      D0CF.. -> eski XLS (xlrd)
    """
    with open(path, "rb") as f:
        head = f.read(8)

    sheets: list[list[tuple]] = []

    if head[:4] == b"\xd0\xcf\x11\xe0":
        # Eski .xls — xlrd bilan o'qiymiz
        import xlrd
        book = xlrd.open_workbook(path)
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    v = cell.value
                    if v == "" or v is None:
                        row.append(None)
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            row.append(str(xlrd.xldate_as_datetime(v, book.datemode)))
                        except Exception:
                            row.append(v)
                    else:
                        row.append(v)
                if any(c is not None for c in row):
                    rows.append(tuple(row))
            if rows:
                sheets.append(rows)
        return sheets

    # XLSX (yoki noma'lum — openpyxl o'zi xato beradi)
    wb = load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        rows = [r for r in ws.iter_rows(values_only=True)
                if r and any(c is not None for c in r)]
        if rows:
            sheets.append(rows)
    return sheets


def _from_filenames(filenames) -> dict:
    """Fayl nomidan korxona nomi va davrni ajratadi.
    Masalan: 'Форма 1  АЛК за 1 кв 2026.xlsx' -> company='АЛК', year=2026, q=1.
    Diqqat: kirillda harf+raqam orasida \\b ishlamaydi, shuning uchun
    lookaround'lar ishlatiladi."""
    company = ""
    year = None
    quarter = None
    if not filenames:
        return {"company": company, "year": year, "quarter": quarter}

    drop_tokens = re.compile(
        r"(?:форм[аы]?|shakl|отчет|hisobot|баланс|balans|квартал|кв[\-\.]?л?|"
        r"chorak|год|yil|за|на|xlsx|xls)", re.IGNORECASE)

    for fn in filenames:
        base = re.sub(r"\.(xlsx?|XLSX?)$", "", str(fn))
        base = base.replace("_", " ").replace("-", " ")
        if year is None:
            m = re.search(r"(?<!\d)(20\d{2})(?!\d)", base)
            if m:
                year = int(m.group(1))
        if quarter is None:
            m = (re.search(r"(?<!\d)([1-4])\s*кв", base, re.IGNORECASE)
                 or re.search(r"кв[а-яё\.\s]{0,6}?([1-4])(?!\d)", base, re.IGNORECASE)
                 or re.search(r"(?<!\d)([1-4])\s*chorak", base, re.IGNORECASE))
            if m:
                quarter = int(m.group(1))
        if not company:
            cleaned = re.sub(r"(?<!\d)20\d{2}(?!\d)", " ", base)
            cleaned = drop_tokens.sub(" ", cleaned)
            cleaned = re.sub(r"\d+", " ", cleaned)
            cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" .-_г")
            # Qoldiq mazmunli bo'lishi shart: kamida bitta 3+ harfli so'z
            if any(len(w) >= 3 for w in cleaned.split()):
                company = cleaned

    return {"company": company, "year": year, "quarter": quarter}


def parse_excel(paths, company: str = "", period: str = "",
                filenames=None) -> FinancialData:
    """Bir yoki bir nechta Excel faylni o'qiydi (Shakl 1 va Shakl 2 alohida
    fayllar bo'lishi mumkin). `paths` — yo'l (str) yoki yo'llar ro'yxati."""
    if isinstance(paths, str):
        paths = [paths]

    found: dict[str, tuple] = {}      # field -> (prev, cur)
    all_rows_text: list[tuple] = []   # kalit so'z zaxirasi uchun

    worksheets = []
    for path in paths:
        worksheets.extend(_load_all_sheets(path))

    if not worksheets:
        raise FormParseError(
            "empty",
            "Yuborilgan Excel fayl(lar) bo'sh yoki o'qib bo'lmadi. "
            "Iltimos, НСБУ formatidagi to'ldirilgan Shakl 1 va Shakl 2 ni yuboring.")

    for rows in worksheets:
        all_rows_text.extend(rows)
        sheet_text = " ".join(
            str(c) for r in rows for c in r if isinstance(c, str)
        ).lower()
        code_col = _detect_code_column(rows)
        codes = _extract_codes(rows, code_col) if code_col is not None else {}
        kind = _classify(sheet_text, set(codes))

        # Sug'urta shaklini aniqlash: 1200 (passiv jami) yoki 930+570 birga
        is_insurance = (1200 in codes) or ({930, 570} <= set(codes))

        maps = []
        if kind in ("form1", "mixed"):
            maps.append(FORM1_INS_CODES if is_insurance else FORM1_CODES)
        if kind in ("form2", "mixed"):
            maps.append(FORM2_CODES)

        for cmap in maps:
            for code, field in cmap.items():
                if field in found:
                    continue
                if code in codes:
                    found[field] = codes[code]

        # Sug'urta rezervlari (720, sof) — uzoq muddatli majburiyatga qo'shamiz,
        # shunda balans ayniyati saqlanadi: 1200 = 570 + 720 + 1190
        if is_insurance and INS_RESERVES_CODE in codes and "long_term_liabilities" in found:
            r_prev, r_cur = codes[INS_RESERVES_CODE]
            l_prev, l_cur = found["long_term_liabilities"]
            found["long_term_liabilities"] = (
                (l_prev or 0) + (r_prev or 0) if (l_prev is not None or r_prev is not None) else None,
                (l_cur or 0) + (r_cur or 0),
            )

    # --- Zaxira: kalit so'z bo'yicha (kod topilmagan maydonlar uchun) ---
    for row in all_rows_text:
        label = " ".join(str(c) for c in row if isinstance(c, str)).lower().strip()
        if not label:
            continue
        nums = [v for v in (_to_float(c) for c in row if _as_code(c) is None)
                if v is not None and v != 0]
        if not nums:
            continue
        cur = nums[-1]
        prev = nums[-2] if len(nums) >= 2 else None
        for field, kws in KEYWORDS.items():
            if field in found:
                continue
            if any(kw in label for kw in kws):
                found[field] = (prev, cur)
                break

    def cur(f):  return found.get(f, (None, 0.0))[1] or 0.0
    def prev(f): return found.get(f, (None, None))[0]

    fd = FinancialData(
        company=company, period=period,
        non_current_assets=cur("non_current_assets"),
        current_assets=cur("current_assets"),
        total_assets=cur("total_assets"),
        inventories=cur("inventories"),
        receivables=cur("receivables"),
        cash=cur("cash"),
        short_term_investments=cur("short_term_investments"),
        equity=cur("equity"),
        long_term_liabilities=cur("long_term_liabilities"),
        current_liabilities=cur("current_liabilities"),
        payables=cur("current_liabilities"),  # alohida bo'lmasa joriy majburiyat
        total_assets_begin=prev("total_assets"),
        current_assets_begin=prev("current_assets"),
        inventories_begin=prev("inventories"),
        receivables_begin=prev("receivables"),
        payables_begin=prev("current_liabilities"),
        equity_begin=prev("equity"),
        cash_begin=prev("cash"),
        non_current_assets_begin=prev("non_current_assets"),
        long_term_liabilities_begin=prev("long_term_liabilities"),
        short_term_investments_begin=prev("short_term_investments"),
        current_liabilities_begin=prev("current_liabilities"),
        revenue=cur("revenue"),
        cogs=cur("cogs"),
        gross_profit=cur("gross_profit"),
        period_expenses=cur("period_expenses"),
        operating_profit=cur("operating_profit"),
        profit_before_tax=cur("profit_before_tax"),
        income_tax=cur("income_tax"),
        net_profit=cur("net_profit"),
        revenue_prev=prev("revenue"),
        cogs_prev=prev("cogs"),
        gross_profit_prev=prev("gross_profit"),
        net_profit_prev=prev("net_profit"),
    )

    # Sarlavhadan korxona nomi, STIR va davrni o'qib qo'shamiz.
    # Ustuvorlik: varaq sarlavhasi > fayl nomi > tashqaridan berilgan qiymat
    hdr = _extract_header(all_rows_text)
    fnm = _from_filenames(filenames)
    fd.company = hdr["company"] or fnm["company"] or company
    fd.stir = hdr["stir"]
    fd.period_year = hdr["year"] if hdr["year"] is not None else fnm["year"]
    fd.period_quarter = (hdr["quarter"] if hdr["quarter"] is not None
                         else fnm["quarter"])

    # Jami qiymatlar НСБУ formulasidan kelib chiqadi, lekin ba'zi fayllarda
    # (emitent hisobotlari) formula natijasi 0 yoki buzuq saqlanadi.
    # Buxgalteriya ayniyati: Jami aktiv = Kapital + Majburiyatlar (passiv).
    # Passiv tomoni odatda ishonchli (480/490/600 + 780 tasdiqlaydi), shuning
    # uchun jami aktivni undan tiklaymiz.
    passive_total = ((fd.equity or 0) + (fd.long_term_liabilities or 0)
                     + (fd.current_liabilities or 0))
    if not fd.total_assets:
        by_assets = (fd.non_current_assets or 0) + (fd.current_assets or 0)
        fd.total_assets = passive_total if passive_total > 0 else by_assets

    # Aktiv tarkibi ziddiyatli bo'lsa (joriy + uzoq ≠ jami, masalan 390 buzuq):
    # uzoq aktivlar (130) ishonchliroq, shuning uchun joriy aktivni
    # ayniyatdan qayta hisoblaymiz: joriy = jami − uzoq.
    if fd.total_assets and fd.non_current_assets is not None:
        breakdown = (fd.current_assets or 0) + fd.non_current_assets
        if abs(breakdown - fd.total_assets) > fd.total_assets * 0.01:
            fd.current_assets = max(fd.total_assets - fd.non_current_assets, 0)
    # Uzoq aktivlar tushib qolgan bo'lsa
    if (not fd.non_current_assets) and fd.total_assets and fd.current_assets:
        fd.non_current_assets = max(fd.total_assets - fd.current_assets, 0)

    # Davr boshi uchun ham jami aktivni passivdan tiklaymiz (pul oqimiga kerak)
    if not fd.total_assets_begin:
        pb = ((fd.equity_begin or 0) + (fd.long_term_liabilities_begin or 0)
              + (fd.current_liabilities_begin or 0))
        ab = (fd.non_current_assets_begin or 0) + (fd.current_assets_begin or 0)
        fd.total_assets_begin = pb if pb > 0 else (ab if ab > 0 else None)

    # Buxgalteriya ayniyati: Majburiyatlar = Aktivlar - Kapital.
    # Joriy yoki uzoq muddatli majburiyat tushmay qolsa — to'ldiramiz.
    total_liab = fd.total_assets - fd.equity
    if not fd.current_liabilities and total_liab > 0:
        fd.current_liabilities = max(total_liab - fd.long_term_liabilities, 0)
        fd.payables = fd.current_liabilities
    if not fd.long_term_liabilities and total_liab > 0 and fd.current_liabilities:
        fd.long_term_liabilities = max(total_liab - fd.current_liabilities, 0)

    # Xuddi shu ayniyat — DAVR BOSHI qiymatlari uchun (pul oqimi tahliliga kerak)
    if (fd.long_term_liabilities_begin is None
            and None not in (fd.total_assets_begin, fd.equity_begin,
                             fd.current_liabilities_begin)):
        fd.long_term_liabilities_begin = max(
            fd.total_assets_begin - fd.equity_begin - fd.current_liabilities_begin, 0)
    if (fd.current_liabilities_begin is None
            and None not in (fd.total_assets_begin, fd.equity_begin,
                             fd.long_term_liabilities_begin)):
        fd.current_liabilities_begin = max(
            fd.total_assets_begin - fd.equity_begin - fd.long_term_liabilities_begin, 0)
        if fd.payables_begin is None:
            fd.payables_begin = fd.current_liabilities_begin

    # current_liabilities qarzsiz firmada 0 bo'lishi mumkin — uni "topildi"
    # deb hisoblaymiz (found da bor yoki balans ayniyatidan kelib chiqqan).
    cl_ok = ("current_liabilities" in found) or (fd.current_liabilities is not None)
    missing = [k for k in ("total_assets", "current_assets", "equity")
               if not getattr(fd, k)]
    if not cl_ok:
        missing.append("current_liabilities")
    if missing:
        # Mijozga tushunarli: maydon nomi + НСБУ qator kodi
        labels = {
            "total_assets":        ("Balans aktivi bo'yicha jami", "400"),
            "current_assets":      ("Joriy aktivlar jami", "390"),
            "equity":              ("O'z kapitali jami", "480"),
            "current_liabilities": ("Joriy majburiyatlar jami", "600"),
        }
        human = [f"«{labels[k][0]}» ({labels[k][1]}-qator)"
                 for k in missing if k in labels]
        raise FormParseError(
            "missing_fields",
            "Shakl 1 (Balans) da quyidagi ko'rsatkichlar topilmadi: "
            + ", ".join(human) + ".",
            fields=human)
    return fd.normalize()


def demo_data() -> FinancialData:
    return FinancialData(
        company='"Namuna" MChJ', period="2025-yil",
        non_current_assets=4200, current_assets=5800, total_assets=10000,
        inventories=2600, receivables=2100, cash=900, short_term_investments=200,
        equity=5500, long_term_liabilities=1500, current_liabilities=3000, payables=2400,
        total_assets_begin=8800, current_assets_begin=5000, inventories_begin=2300,
        receivables_begin=1700, payables_begin=2100, equity_begin=4800,
        revenue=18000, cogs=12600, gross_profit=5400, period_expenses=3400,
        operating_profit=2000, profit_before_tax=1700, income_tax=255, net_profit=1445,
        revenue_prev=15500, cogs_prev=11200, gross_profit_prev=4300, net_profit_prev=1050,
    ).normalize()